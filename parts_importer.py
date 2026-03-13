#!/usr/bin/env python3
"""Import parts data from work-ticket Excel sheets into PostgreSQL.

Parses legacy hand-written work tickets to extract parts mapping
(category, product_code, part_name, material, processes).
Deduplicates by (category, product_code, part_name, material), keeping the longest process list.

Requires DATABASE_URL env var (e.g. postgresql://cck:cck@localhost:5432/cck).
"""

import argparse
import os
import openpyxl
import pandas as pd
import psycopg2
import re
import sys
import time
from typing import Dict, List, Optional, Tuple


class WorkTicketParser:
    def __init__(self, filename: str):
        self.filename = filename
        self.header_pattern = "四川成都空分配套阀门有限公司"

    def clean_material(self, raw: Optional[str]) -> Optional[str]:
        if not raw:
            return None
        cleaned = raw.replace('"', '').replace('\u201c', '').replace('\u201d', '').strip()
        return re.sub(r'\s+', ' ', cleaned)

    def populate_db(self, dfs):
        """Write parsed data into PostgreSQL database"""
        database_url = os.environ.get('DATABASE_URL')
        if not database_url:
            print("ERROR: DATABASE_URL env var is required", file=sys.stderr)
            sys.exit(1)

        conn = psycopg2.connect(database_url)
        cur = conn.cursor()
        cur.execute('DROP TABLE IF EXISTS work_tickets')
        cur.execute('''CREATE TABLE work_tickets (
            id SERIAL PRIMARY KEY,
            category TEXT NOT NULL,
            product_code TEXT NOT NULL,
            part_name TEXT,
            material TEXT,
            processes TEXT,
            process_count INTEGER DEFAULT 0
        )''')
        cur.execute('CREATE INDEX idx_wt_category ON work_tickets(category)')
        cur.execute('CREATE INDEX idx_wt_product_code ON work_tickets(product_code)')
        for df in dfs:
            for _, row in df.iterrows():
                cur.execute(
                    'INSERT INTO work_tickets (category, product_code, part_name, material, processes, process_count) VALUES (%s, %s, %s, %s, %s, %s)',
                    (row['category'], row['product_code'], row['part_name'],
                     row['material'], row['processes'], row['process_count'])
                )
        conn.commit()
        cur.execute('SELECT COUNT(*) FROM work_tickets')
        total = cur.fetchone()[0]
        cur.close()
        conn.close()
        print(f"\nPopulated PostgreSQL with {total} records")

    def parse_product_code_range(self, code_string: str) -> List[Tuple[str, str]]:
        """Parse product codes and extract both code and part name"""
        results = []

        parts = code_string.split('.')

        if len(parts) == 2 and not re.match(r'^\d+/\d+', parts[1]):
            product_code = parts[0]
            part_name = parts[1]
        elif len(parts) == 3 and re.match(r'^\d+/\d+', parts[1]):
            product_code = f"{parts[0]}.{parts[1]}"
            part_name = parts[2] if len(parts[2]) > 0 else None
        elif len(parts) == 2 and re.match(r'^\d+/\d+', parts[1]):
            product_code = code_string
            part_name = None
        else:
            product_code = code_string
            part_name = None

        if '～' in product_code or '~' in product_code:
            match = re.match(r'T(\d+)[～~](\d+)(.+)', product_code)
            if match:
                start = int(match.group(1))
                end = int(match.group(2))
                suffix = match.group(3)

                for num in range(start, end + 1):
                    results.append((f"T{num}{suffix}", part_name))
            else:
                results.append((product_code, part_name))
        else:
            results.append((product_code, part_name))

        return results

    def extract_ticket_info(self, sheet, start_row: int, end_row: int) -> Dict:
        """Extract specific information from a work ticket"""
        info = {
            'product_code': None,
            'part_name': None,
            'material': None,
            'processes': [],
            'raw_product_string': None
        }

        content = {}
        for row in range(start_row, min(end_row + 1, start_row + 50)):
            row_data = []
            for col in range(1, 21):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    row_data.append({
                        'col': col,
                        'value': str(cell.value).strip()
                    })
            if row_data:
                content[row] = row_data

        product_row = start_row + 3
        header_labels = {"产品代号", "零件名称", "代用"}
        if product_row in content:
            for i, item in enumerate(content[product_row]):
                if item['value'] == "产品代号" and i + 1 < len(content[product_row]):
                    next_val = content[product_row][i + 1]['value']
                    if next_val not in header_labels:
                        info['raw_product_string'] = next_val
                elif item['value'] == "零件名称" and i + 1 < len(content[product_row]):
                    next_val = content[product_row][i + 1]['value']
                    if next_val not in header_labels:
                        info['part_name'] = next_val

        material_header_row = start_row + 4
        material_data_row = start_row + 5

        if material_header_row in content and material_data_row in content:
            material_col = None
            for item in content[material_header_row]:
                if item['value'] == "材质及编号":
                    material_col = item['col']
                    break

            if material_col:
                for item in content[material_data_row]:
                    if item['col'] == material_col:
                        info['material'] = self.clean_material(item['value'])
                        break

        if not info['material']:
            for row_num, row_data in content.items():
                for item in row_data:
                    val = item['value']
                    if ('06Cr19Ni10' in val or '5083' in val or '板料' in val or
                        '不锈钢' in val or 'Q235' in val or '1Cr18Ni9Ti' in val):
                        if '注意' not in val and len(val) < 50:
                            info['material'] = self.clean_material(val)
                            break
                if info['material']:
                    break

        for row_num in sorted(content.keys()):
            row_data = content[row_num]

            if len(row_data) >= 2:
                first_val = row_data[0]['value']
                if re.match(r'^\d+\.?\d*$', first_val):
                    if len(row_data) > 1:
                        process_name = row_data[1]['value']
                        if (not process_name.startswith('注意') and
                            not process_name.startswith('见') and
                            len(process_name) < 30):
                            process_keywords = ['下料', '车', '清洗', '焊接', '钳', '校圆',
                                              '卷筒', '钻', '铣', '磨', '镗', '探伤', '线切割',
                                              '钳工', '装配', '检验', '外协', '研磨']
                            if any(keyword in process_name for keyword in process_keywords):
                                info['processes'].append(process_name)

        return info

    def handle_duplicates(self, tickets: List[Dict]) -> List[Dict]:
        """Deduplicate by (category, product_code, part_name, material), keeping longest process list"""
        best = {}
        for ticket in tickets:
            key = (
                ticket['category'],
                ticket['product_code'],
                ticket['part_name'],
                ticket['material'],
            )
            process_list = ticket['processes'].split(' → ') if ticket['processes'] else []
            ticket['process_count'] = len(process_list)

            if key not in best or ticket['process_count'] > best[key]['process_count']:
                best[key] = ticket

        return list(best.values())

    def parse_sheet(self, sheet_name: str, category: str) -> pd.DataFrame:
        """Parse all work tickets in a sheet and return DataFrame"""
        print(f"\n{'='*60}")
        print(f"Processing Sheet {sheet_name}: {category}")
        print('='*60)

        print("Loading Excel file (this may take 20-30 seconds)...", end="", flush=True)
        start_time = time.time()
        wb = openpyxl.load_workbook(self.filename, data_only=True)
        sheet = wb[sheet_name]
        print(f" done ({time.time() - start_time:.1f}s)")

        print(f"Sheet has {sheet.max_row} rows")
        print("Searching for work tickets...")

        headers_found = []
        for row in range(1, min(sheet.max_row + 1, 10000)):
            cell = sheet.cell(row=row, column=1)
            if cell.value and self.header_pattern in str(cell.value):
                headers_found.append(row)
                if len(headers_found) % 50 == 0:
                    print(f"  Found {len(headers_found)} tickets so far...")

        print(f"Found {len(headers_found)} work tickets total")

        print("Extracting ticket data...")
        tickets = []
        for i, start_row in enumerate(headers_found):
            if (i + 1) % 50 == 0:
                print(f"  Processed {i+1}/{len(headers_found)} tickets...")

            end_row = headers_found[i + 1] - 1 if i + 1 < len(headers_found) else start_row + 30

            info = self.extract_ticket_info(sheet, start_row, end_row)

            if info['raw_product_string']:
                code_results = self.parse_product_code_range(info['raw_product_string'])

                for product_code, extracted_part_name in code_results:
                    ticket = {
                        'category': category,
                        'product_code': product_code,
                        'part_name': info['part_name'] or extracted_part_name,
                        'material': info['material'],
                        'processes': ' → '.join(info['processes']) if info['processes'] else '',
                    }
                    tickets.append(ticket)

        wb.close()

        print("Processing duplicates...")
        tickets = self.handle_duplicates(tickets)

        df = pd.DataFrame(tickets)

        columns = ['category', 'product_code', 'part_name',
                  'material', 'processes', 'process_count']
        df = df[columns]

        print(f"Extracted {len(df)} records")

        return df

    def parse_all_sheets(self, sheet_numbers: List[int] = None):
        """Parse specified sheets"""
        categories = {
            "1": "冷箱阀DL",
            "2": "冷箱阀TJ",
            "3": "冷箱阀TC",
            "4": "冷箱阀TD",
            "5": "冷箱阀TN",
            "6": "冷箱阀TE",
            "7": "冷箱阀TG",
            "8": "冷箱阀TF",
            "9": "冷箱阀TH",
            "10": "冷箱阀TI",
            "11": "冷箱阀TK",
            "12": "冷箱阀TL"
        }

        if sheet_numbers:
            sheets_to_parse = [str(s) for s in sheet_numbers]
        else:
            sheets_to_parse = list(categories.keys())

        print(f"\nWill parse {len(sheets_to_parse)} sheet(s)")

        all_dfs = []
        for sheet_name in sheets_to_parse:
            category = categories.get(sheet_name, f"Sheet_{sheet_name}")
            df = self.parse_sheet(sheet_name, category)
            all_dfs.append(df)

        return all_dfs


def main():
    arg_parser = argparse.ArgumentParser(description='Import parts data from work-ticket Excel sheets into PostgreSQL')
    arg_parser.add_argument('--file', type=str, default='gp_no_passwords.xlsx',
                            help='Excel file to parse (default: gp_no_passwords.xlsx)')
    arg_parser.add_argument('--sheet', type=int, choices=range(1, 13), metavar='N',
                            help='Sheet number (1-12). Omit to parse all sheets.')
    arg_parser.add_argument('--populate-db', action='store_true',
                            help='Write parsed data into PostgreSQL (requires DATABASE_URL)')
    args = arg_parser.parse_args()

    sheet_numbers = [args.sheet] if args.sheet else None
    if args.sheet:
        print(f"Testing with sheet {args.sheet} only")
    else:
        print("Parsing all 12 sheets")

    start_time = time.time()
    parser = WorkTicketParser(args.file)

    dfs = parser.parse_all_sheets(sheet_numbers)

    if args.populate_db:
        parser.populate_db(dfs)

    print(f"\n{'='*60}")
    print("FINAL SUMMARY")
    print('='*60)

    total_records = sum(len(df) for df in dfs)
    print(f"Total records extracted: {total_records}")
    print(f"Total time: {time.time() - start_time:.1f} seconds")

    for df in dfs:
        if not df.empty:
            category = df['category'].iloc[0]
            print(f"\n{category}:")
            print(f"  Total records: {len(df)}")
            print(f"  Unique products: {df['product_code'].nunique()}")

            wrong_mat = df[df['material'] == '定额']
            if not wrong_mat.empty:
                print(f"  Items with '定额' as material: {len(wrong_mat)}")

if __name__ == "__main__":
    main()
